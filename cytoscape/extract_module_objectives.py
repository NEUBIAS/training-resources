"""
Extract module learning objectives from Markdown files with YAML front-matter
and write them into an Excel workbook.

This script scans the repository's _modules directory for Markdown (*.md)
files that contain YAML front-matter. For each file it attempts to parse the
YAML metadata and, if present and not excluded by tags, collects the module
title, learning objectives, tags and filename. The collected modules are
sorted according to tag-based rules and written to an Excel workbook with a
hyperlink to the module page.

Functions
---------
extract_metadata_from_md(md_path)
    Parse YAML front-matter from the Markdown file at md_path.

    Parameters
    ----------
    md_path : str or os.PathLike
        Filesystem path to a Markdown file expected to start with '---' and
        contain YAML between the opening and closing '---' markers.

    Returns
    -------
    dict or None
        On success, returns a dictionary with keys:
            - 'title' (str): title from YAML or filename fallback
            - 'objectives' (str): objectives as a single string (each item
              normalized to end with a period); empty string if missing
            - 'tags' (list[str]): list of lower-cased tags
            - 'filename' (str): filename without extension
        Returns None if the file has no YAML front-matter, if the YAML is
        empty, or if the module is explicitly excluded (see Filtering).
        If YAML parsing fails, the exception is caught, an error is printed
        and None is returned.

    Notes
    -----
    - Accepts 'tags' as either a string or a list; normalizes to a list.
    - Accepts 'objectives' as either a string or a list; normalizes to a
      single sentence-per-item string.
    - Files with tags containing any of ['draft', 'outdated', 'course']
      (case-insensitive) are excluded.

sort_key(module)
    Compute a sorting key for a module dictionary.

    Parameters
    ----------
    module : dict
        Module dictionary as returned by extract_metadata_from_md.

    Returns
    -------
    tuple
        A tuple used for sorting:
            - First element is an integer priority: 1 (default), 2 ('workflow'),
              3 ('scripting').
            - Second element is the module title lower-cased for
              alphabetical ordering within the priority.

    Discover Markdown module files, extract metadata, sort modules and write
    an Excel workbook named "module_objectives.xlsx" containing the results.

    Behavior and side effects
    -------------------------
    - The script assumes the _modules directory is located one level up from
      the script's directory (../_modules).
    - All *.md files in that directory are processed.
    - Uses openpyxl to create an Excel workbook with a single sheet "Objectives".
      Columns written:
        1. "Module title" — the title text with a hyperlink to the module page
           constructed as:
           https://neubias.github.io/training-resources/{filename}
        2. "Learning objectives" — the normalized objectives string.
    - Saves the workbook to "module_objectives.xlsx" in the current working
      directory.

Dependencies
------------
- PyYAML (yaml) for parsing YAML front-matter.
- openpyxl for writing Excel files.

Example
-------
If a module file contains YAML like:

---
title: Example Module
tags: [analysis, workflow]
objectives:
  - Understand X
  - Learn Y
---

extract_metadata_from_md will return a dict with:
    {
      'title': 'Example Module',
      'objectives': 'Understand X. Learn Y.',
      'tags': ['analysis', 'workflow'],
      'filename': 'example-module'

Notes
-----
- Parsing errors are printed to stdout; the script continues processing other
  files rather than raising.
- The output file name and URL base are currently hard-coded.
"""
import os
import glob
import yaml
import openpyxl

def extract_metadata_from_md(md_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    if not lines or not lines[0].strip().startswith('---'):
        return None
    yaml_lines = []
    for line in lines[1:]:
        if line.strip().startswith('---'):
            break
        yaml_lines.append(line)
    try:
        meta = yaml.safe_load(''.join(yaml_lines))
        if not meta:
            return None
        tags = meta.get('tags', [])
        if isinstance(tags, str):
            tags = [tags]
        # Exclude drafts and outdated
        if any(tag.lower() in ['draft', 'outdated', 'course'] for tag in tags):
            return None
        title = meta.get('title', os.path.basename(md_path))
        objectives = meta.get('objectives', None)
        if isinstance(objectives, list):
            objectives_str = ' '.join(obj.strip().rstrip('.') + '.' for obj in objectives)
        elif isinstance(objectives, str):
            objectives_str = objectives.strip().rstrip('.') + '.'
        else:
            objectives_str = ''
        filename = os.path.splitext(os.path.basename(md_path))[0]
        return {
            'title': title,
            'objectives': objectives_str,
            'tags': [t.lower() for t in tags],
            'filename': filename
        }
    except Exception as e:
        print(f"Error parsing {md_path}: {e}")
        return None

def sort_key(module):
    tags = module['tags']
    if 'workflow' in tags:
        return (2, module['title'].lower())
    if 'scripting' in tags:
        return (3, module['title'].lower())
    return (1, module['title'].lower())

def main():
    modules_dir = os.path.join(os.path.dirname(__file__), '..', '_modules')
    md_files = glob.glob(os.path.join(modules_dir, '*.md'))
    modules = []
    for md_file in md_files:
        meta = extract_metadata_from_md(md_file)
        if meta:
            modules.append(meta)
    modules.sort(key=sort_key)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Objectives"
    ws.append(["Module title", "Learning objectives"])
    for module in modules:
        url = f"https://neubias.github.io/training-resources/{module['filename']}"
        cell = ws.cell(row=ws.max_row + 1, column=1, value=module['title'])
        cell.hyperlink = url
        cell.style = "Hyperlink"
        ws.cell(row=cell.row, column=2, value=module['objectives'])
    wb.save("module_objectives.xlsx")

if __name__ == "__main__":
    main()